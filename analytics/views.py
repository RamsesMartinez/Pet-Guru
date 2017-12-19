from datetime import datetime

from django.views.generic import TemplateView
from django.shortcuts import render
from django.http import HttpResponse

# libreria para el manejo de archivos xls
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Models
from users.models import User

class ReporteUsuarios(TemplateView):
    """
    Vista basada en clases que retorna un archivo de reporte de formato xls (Excel)
        HttpResponse de content_type='application/ms-excel'
    Para obtener archivo, debe ser llamado desde petición get
    """
    def get(self, request, *args, **kwargs):
        """
        Funcionamiento:
            Para la generación de los campos, unicamente se deben cargar los modelos
            y sus campos respectivos, así como sus inner joins que se requieren para 
            mostrarse en el excel.
        """
        # Se cargan los modelos para insertar la información
        users = User.objects.all()
        # Se el objeto xls
        workbook = Workbook()
        # Contador para iniciar desde la fila 4 a colocar la información de los modelos   
        count = 4
        # Activa la hoja de excel 1 para trabajar
        ws1 = workbook.active
        
        # Inserta contenido a celdas
        ws1.title = "Reporte de usuarios"
        ws1['A1'] = 'Reporte general de usuarios PETGURU'
        
        # Aplica estilos a las celdas
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('A1:D1')
        
        # Se definen los encabezados de las columnas
        ws1['A3'] = 'ID'
        ws1['B3'] = 'Rol'
        ws1['C3'] = 'Especialidad'

        for user in users:
            # Itera celdas y añade contenido
            ws1.cell(row=count, column=1, value=user.id)
            ws1.cell(row=count, column=2, value=user.rol)
            if not user.speciality:
                ws1.cell(row=count, column=3, value='sin especialidad')
            else:
                ws1.cell(row=count, column=3, value=user.speciality)
            count += 1
            
        file_name = 'Reporte_General_De_Usuarios_{0}.xlsx'.format(datetime.now().strftime("%I-%M%p_%d-%m-%Y"))
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename={0}'.format(file_name)
        response['Content-Disposition'] = content
        workbook.save(response)

        return response
